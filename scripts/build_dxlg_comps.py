"""Build DXLG comparable company analysis (specialty apparel small-cap peer set)."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

OUT = "/home/user/Claude-code-work/output/DXLG_comps.xlsx"

# All $ in millions. Net debt: negative = net cash. Growth as decimal.
# Each ref tags the source 10-K cited per cell.
# NOTE: Most peers have Feb fiscal year-ends → reasonably tight period alignment.
DATA = {
    "DXLG": dict(
        revenue=467, growth=-0.080, gross_profit=214, adj_ebitda=38, fcf=15,
        share_price=2.25, shares=63, net_debt=-45, eps=0.13,
        ref="Destination XL Group FY24 10-K (year ended Feb 1, 2025); Adjusted EBITDA per non-GAAP recon",
        period="FY24 (LTM through Feb 1, 2025)",
    ),
    "BKE": dict(
        revenue=1219, growth=-0.026, gross_profit=555, adj_ebitda=270, fcf=155,
        share_price=45.00, shares=50, net_debt=-280, eps=4.05,
        ref="The Buckle Inc FY24 10-K (year ended Feb 1, 2025); pays large recurring special dividends",
        period="FY24 (LTM through Feb 1, 2025)",
    ),
    "CATO": dict(
        revenue=675, growth=-0.070, gross_profit=240, adj_ebitda=15, fcf=5,
        share_price=4.00, shares=21, net_debt=-50, eps=-0.50,
        ref="Cato Corp FY24 10-K (year ended Feb 1, 2025); EPS reflects GAAP loss",
        period="FY24 (LTM through Feb 1, 2025)",
    ),
    "CTRN": dict(
        revenue=745, growth=0.010, gross_profit=285, adj_ebitda=15, fcf=5,
        share_price=25.00, shares=8.4, net_debt=-60, eps=-1.80,
        ref="Citi Trends FY24 10-K (year ended Feb 1, 2025); EPS reflects GAAP loss; EBITDA depressed by store-level deleverage",
        period="FY24 (LTM through Feb 1, 2025)",
    ),
    "TLYS": dict(
        revenue=585, growth=-0.090, gross_profit=170, adj_ebitda=-5, fcf=-15,
        share_price=3.50, shares=30, net_debt=-30, eps=-1.15,
        ref="Tilly's Inc FY24 10-K (year ended Feb 1, 2025); negative Adj EBITDA — EV/EBITDA shown as n/m",
        period="FY24 (LTM through Feb 1, 2025)",
    ),
    "ZUMZ": dict(
        revenue=885, growth=0.015, gross_profit=295, adj_ebitda=30, fcf=25,
        share_price=15.00, shares=19.5, net_debt=-140, eps=0.10,
        ref="Zumiez Inc FY24 10-K (year ended Feb 1, 2025); large net cash position",
        period="FY24 (LTM through Feb 1, 2025)",
    ),
    "GCO": dict(
        revenue=2318, growth=0.010, gross_profit=1100, adj_ebitda=60, fcf=25,
        share_price=25.00, shares=10.8, net_debt=180, eps=-0.90,
        ref="Genesco Inc FY25 10-K (year ended Feb 1, 2025); Journeys/Schuh/Johnston & Murphy; carries net debt",
        period="FY25 (LTM through Feb 1, 2025)",
    ),
    "DBI": dict(
        revenue=3011, growth=-0.030, gross_profit=910, adj_ebitda=135, fcf=50,
        share_price=5.00, shares=55, net_debt=200, eps=0.10,
        ref="Designer Brands Inc FY24 10-K (year ended Feb 1, 2025); DSW parent",
        period="FY24 (LTM through Feb 1, 2025)",
    ),
    "CURV": dict(
        revenue=1089, growth=-0.040, gross_profit=420, adj_ebitda=110, fcf=50,
        share_price=5.00, shares=105, net_debt=300, eps=0.25,
        ref="Torrid Holdings FY24 10-K (year ended Feb 1, 2025); plus-size women's apparel; carries LBO-era term loan",
        period="FY24 (LTM through Feb 1, 2025)",
    ),
}

NAVY = "17365D"
LIGHT_BLUE = "D9E1F2"
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"
INPUT_BLUE = "0070C0"

FONT_NAME = "Times New Roman"

PEERS = [
    ("Destination XL Group", "DXLG"),
    ("The Buckle", "BKE"),
    ("Cato Corp", "CATO"),
    ("Citi Trends", "CTRN"),
    ("Tilly's", "TLYS"),
    ("Zumiez", "ZUMZ"),
    ("Genesco", "GCO"),
    ("Designer Brands", "DBI"),
    ("Torrid Holdings", "CURV"),
]

OPS_COLS = [
    ("Company", 24, "left"),
    ("Revenue (LTM, $M)", 18, "center"),
    ("Rev Growth (YoY)", 16, "center"),
    ("Gross Profit ($M)", 18, "center"),
    ("Gross Margin", 14, "center"),
    ("EBITDA ($M)", 16, "center"),
    ("EBITDA Margin", 14, "center"),
    ("FCF ($M)", 14, "center"),
    ("FCF Margin", 12, "center"),
]

VAL_COLS = [
    ("Company", 24, "left"),
    ("Share Price ($)", 14, "center"),
    ("Diluted Shares (M)", 16, "center"),
    ("Market Cap ($M)", 16, "center"),
    ("Net Debt ($M)", 14, "center"),
    ("Enterprise Value ($M)", 18, "center"),
    ("EV / Revenue", 14, "center"),
    ("EV / EBITDA", 14, "center"),
    ("P / E (LTM)", 14, "center"),
    ("EPS (LTM, $)", 14, "center"),
]

STAT_ROWS = [
    ("Maximum", "MAX"),
    ("75th Percentile", "QUARTILE_3"),
    ("Median", "MEDIAN"),
    ("25th Percentile", "QUARTILE_1"),
    ("Minimum", "MIN"),
]


def style_cell(cell, *, bold=False, color=None, fill=None, align="left", size=11, italic=False, number_format=None):
    cell.font = Font(name=FONT_NAME, size=size, bold=bold, italic=italic, color=color or "000000")
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if number_format:
        cell.number_format = number_format


def section_header(ws, row, span_cols, text):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    for col in range(1, span_cols + 1):
        c = ws.cell(row=row, column=col)
        style_cell(c, bold=True, color=WHITE, fill=NAVY, align="center", size=12)
    ws.row_dimensions[row].height = 22


def column_headers(ws, row, cols):
    for i, (name, _w, align) in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=name)
        style_cell(c, bold=True, fill=LIGHT_BLUE, align="center", size=11)
    ws.row_dimensions[row].height = 30


def set_col_widths(ws, cols):
    for i, (_n, w, _a) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_company_names(ws, start_row, cols):
    for i, (name, ticker) in enumerate(PEERS):
        r = start_row + i
        c = ws.cell(row=r, column=1, value=f"{name} ({ticker})")
        # Bold DXLG row to highlight subject company
        style_cell(c, bold=(ticker == "DXLG"), align="left", size=11)
        for col_idx in range(2, len(cols) + 1):
            cc = ws.cell(row=r, column=col_idx)
            style_cell(cc, bold=(ticker == "DXLG"), align="center", size=11)
        ws.row_dimensions[r].height = 20


def set_input(ws, row, col, value, source_ref, field_label, bold=False):
    """Set an input cell value and attach a source-citation comment."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT_NAME, size=11, color=INPUT_BLUE, bold=bold)
    c.comment = Comment(
        f"{field_label}\nSource: {source_ref}\nNote: Figure from training-data recall (Jan 2026 cutoff). VERIFY against the cited filing on SEC EDGAR before publishing or trading on this output.",
        "Comps",
    )


def stats_block(ws, data_start_row, data_end_row, stats_start_row, cols_to_stat, n_cols):
    for j, (label, fn) in enumerate(STAT_ROWS):
        r = stats_start_row + j
        c = ws.cell(row=r, column=1, value=label)
        style_cell(c, bold=True, fill=LIGHT_GREY, align="left", size=11, italic=True)
        for col_idx in range(2, n_cols + 1):
            cc = ws.cell(row=r, column=col_idx)
            style_cell(cc, fill=LIGHT_GREY, align="center", size=11)
            if col_idx in cols_to_stat:
                col_letter = get_column_letter(col_idx)
                rng = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"
                if fn == "MAX":
                    formula = f"=MAX({rng})"
                elif fn == "MIN":
                    formula = f"=MIN({rng})"
                elif fn == "MEDIAN":
                    formula = f"=MEDIAN({rng})"
                elif fn == "QUARTILE_1":
                    formula = f"=QUARTILE({rng},1)"
                elif fn == "QUARTILE_3":
                    formula = f"=QUARTILE({rng},3)"
                cc.value = formula
        ws.row_dimensions[r].height = 20


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Comps"
    ws.sheet_view.showGridLines = False

    set_col_widths(ws, OPS_COLS)
    if len(VAL_COLS) > len(OPS_COLS):
        for i in range(len(OPS_COLS) + 1, len(VAL_COLS) + 1):
            ws.column_dimensions[get_column_letter(i)].width = VAL_COLS[i - 1][1]

    max_cols = max(len(OPS_COLS), len(VAL_COLS))

    # ----- Header block (rows 1-3) -----
    title = "SPECIALTY APPAREL RETAIL — DXLG COMPARABLE COMPANY ANALYSIS"
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
    for col in range(1, max_cols + 1):
        style_cell(ws.cell(row=1, column=col), bold=True, color=WHITE, fill=NAVY, align="center", size=14)
    ws.row_dimensions[1].height = 28

    sub = " • ".join([f"{n} ({t})" for n, t in PEERS])
    ws.cell(row=2, column=1, value=sub)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_cols)
    style_cell(ws.cell(row=2, column=1), align="center", size=10, italic=True)
    ws.row_dimensions[2].height = 18

    ctx = "As of LTM through each company's most-recent reported quarter (most ~Feb 1, 2025) | All figures in USD Millions except per-share amounts and ratios"
    ws.cell(row=3, column=1, value=ctx)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_cols)
    style_cell(ws.cell(row=3, column=1), align="center", size=10, italic=True)
    ws.row_dimensions[3].height = 18

    # ----- Section 1: Operating Metrics -----
    section_header(ws, 5, len(OPS_COLS), "OPERATING STATISTICS & FINANCIAL METRICS")
    column_headers(ws, 6, OPS_COLS)

    OPS_DATA_START = 7
    OPS_DATA_END = OPS_DATA_START + len(PEERS) - 1
    write_company_names(ws, OPS_DATA_START, OPS_COLS)

    for i, (name, ticker) in enumerate(PEERS):
        r = OPS_DATA_START + i
        d = DATA[ticker]
        ref = d["ref"]
        bold = (ticker == "DXLG")
        set_input(ws, r, 2, d["revenue"], ref, f"{ticker} — Revenue (LTM, $M), {d['period']}", bold=bold)
        set_input(ws, r, 3, d["growth"], ref, f"{ticker} — Revenue YoY growth, {d['period']}", bold=bold)
        set_input(ws, r, 4, d["gross_profit"], ref, f"{ticker} — Gross Profit ($M), {d['period']}", bold=bold)
        set_input(ws, r, 6, d["adj_ebitda"], ref, f"{ticker} — Adjusted (non-GAAP) EBITDA ($M), {d['period']}", bold=bold)
        set_input(ws, r, 8, d["fcf"], ref, f"{ticker} — Free Cash Flow ($M), {d['period']}", bold=bold)

    # Formula columns: E=GM, G=EBITDA-M, I=FCF-M (all use IFERROR for safety)
    for i, (_n, ticker) in enumerate(PEERS):
        r = OPS_DATA_START + i
        bold = (ticker == "DXLG")
        c = ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/B{r},\"\")")
        style_cell(c, align="center", number_format="0.0%", bold=bold)
        c = ws.cell(row=r, column=7, value=f"=IFERROR(F{r}/B{r},\"\")")
        style_cell(c, align="center", number_format="0.0%", bold=bold)
        c = ws.cell(row=r, column=9, value=f"=IFERROR(H{r}/B{r},\"\")")
        style_cell(c, align="center", number_format="0.0%", bold=bold)

    for i in range(len(PEERS)):
        r = OPS_DATA_START + i
        for col_idx in [2, 4, 6, 8]:
            ws.cell(row=r, column=col_idx).number_format = "#,##0;(#,##0)"
        ws.cell(row=r, column=3).number_format = "0.0%"

    OPS_STATS_START = OPS_DATA_END + 2
    OPS_STATS_COLS = [3, 5, 7, 9]
    stats_block(ws, OPS_DATA_START, OPS_DATA_END, OPS_STATS_START, OPS_STATS_COLS, len(OPS_COLS))
    for j in range(len(STAT_ROWS)):
        r = OPS_STATS_START + j
        for col_idx in OPS_STATS_COLS:
            ws.cell(row=r, column=col_idx).number_format = "0.0%"

    # ----- Section 2: Valuation Multiples -----
    VAL_HEADER_ROW = OPS_STATS_START + len(STAT_ROWS) + 2
    section_header(ws, VAL_HEADER_ROW, len(VAL_COLS), "VALUATION MULTIPLES & INVESTMENT METRICS")
    column_headers(ws, VAL_HEADER_ROW + 1, VAL_COLS)

    VAL_DATA_START = VAL_HEADER_ROW + 2
    VAL_DATA_END = VAL_DATA_START + len(PEERS) - 1
    write_company_names(ws, VAL_DATA_START, VAL_COLS)

    for i, (name, ticker) in enumerate(PEERS):
        r = VAL_DATA_START + i
        d = DATA[ticker]
        ref = d["ref"]
        bold = (ticker == "DXLG")
        set_input(ws, r, 2, d["share_price"],
                  "Approximate close as of 2026-05-22; placeholder — replace with verified market data",
                  f"{ticker} — Share Price ($), illustrative as-of 2026-05-22", bold=bold)
        set_input(ws, r, 3, d["shares"], ref,
                  f"{ticker} — Diluted Shares Outstanding (M), per most recent 10-Q cover page", bold=bold)
        set_input(ws, r, 5, d["net_debt"], ref,
                  f"{ticker} — Net Debt ($M, negative = net cash) = Total Debt - Cash & ST Investments, most recent b/s", bold=bold)
        set_input(ws, r, 10, d["eps"], ref,
                  f"{ticker} — Diluted EPS (LTM, $), {d['period']}", bold=bold)

    # Formulas — handles negative EBITDA / negative EPS gracefully ("n/m" → "")
    for i, (_n, ticker) in enumerate(PEERS):
        vr = VAL_DATA_START + i
        ops_r = OPS_DATA_START + i
        bold = (ticker == "DXLG")

        # Market Cap = Share Price * Diluted Shares
        c = ws.cell(row=vr, column=4, value=f"=B{vr}*C{vr}")
        style_cell(c, align="center", number_format="#,##0", bold=bold)

        # EV = Market Cap + Net Debt
        c = ws.cell(row=vr, column=6, value=f"=D{vr}+E{vr}")
        style_cell(c, align="center", number_format="#,##0", bold=bold)

        # EV/Revenue (revenue always positive — IFERROR for safety)
        c = ws.cell(row=vr, column=7, value=f"=IFERROR(F{vr}/B{ops_r},\"\")")
        style_cell(c, align="center", number_format='0.0"x"', bold=bold)

        # EV/EBITDA — explicit guard against zero/negative EBITDA (returns blank, excluded from stats)
        c = ws.cell(row=vr, column=8, value=f"=IF(F{ops_r}>0,F{vr}/F{ops_r},\"\")")
        style_cell(c, align="center", number_format='0.0"x"', bold=bold)

        # P/E — explicit guard against negative EPS
        c = ws.cell(row=vr, column=9, value=f"=IF(J{vr}>0,B{vr}/J{vr},\"\")")
        style_cell(c, align="center", number_format='0.0"x"', bold=bold)

        ws.cell(row=vr, column=2).number_format = "$#,##0.00"
        ws.cell(row=vr, column=3).number_format = "#,##0.0"
        ws.cell(row=vr, column=5).number_format = "#,##0;(#,##0)"
        ws.cell(row=vr, column=10).number_format = "$#,##0.00;($#,##0.00)"

    VAL_STATS_START = VAL_DATA_END + 2
    VAL_STATS_COLS = [7, 8, 9]
    stats_block(ws, VAL_DATA_START, VAL_DATA_END, VAL_STATS_START, VAL_STATS_COLS, len(VAL_COLS))
    for j in range(len(STAT_ROWS)):
        r = VAL_STATS_START + j
        for col_idx in VAL_STATS_COLS:
            ws.cell(row=r, column=col_idx).number_format = '0.0"x"'

    # ----- Section 3: Notes & Methodology -----
    NOTES_ROW = VAL_STATS_START + len(STAT_ROWS) + 2
    section_header(ws, NOTES_ROW, len(VAL_COLS), "NOTES & METHODOLOGY")

    notes = [
        ("⚠ VERIFY BEFORE USE", "All input figures (blue cells) were populated from training-data recall (Jan 2026 cutoff) citing each company's most recent SEC filing. Every blue cell has a comment with the source filing. VERIFY each figure against the actual EDGAR filing before publishing, trading, or presenting. Share prices are illustrative placeholders as of 2026-05-22 and MUST be refreshed against live market data."),
        ("Subject company", "DXLG (Destination XL Group) is the subject of this analysis — its row is bolded across both blocks. DXLG operates ~290 Destination XL and Casual Male XL stores serving the US big & tall men's apparel niche. Micro-cap (~$140M market cap on these assumptions), no direct public pure-play competitor."),
        ("Peer selection rationale", "Selected 8 specialty apparel retailers as the closest liquid public comparison set. BKE/CATO/ZUMZ/CTRN/TLYS are small-cap mall/strip-mall specialty apparel. GCO/DBI are footwear-led specialty retail (Journeys, DSW). CURV (Torrid) is the only true niche/category-specialty parallel (plus-size women). All operate at meaningfully different scales than DXLG; reader should weight comparability accordingly."),
        ("Reporting periods", "Most peers have Feb fiscal year-ends, giving tight period alignment (FY24/FY25 ended ~Feb 1, 2025). DXLG, BKE, CATO, CTRN, TLYS, ZUMZ, DBI, CURV all align on Feb 2025 year-ends; GCO uses FY25 nomenclature for the same period."),
        ("EBITDA definition", "Adjusted (non-GAAP) EBITDA per each company's investor presentation reconciliation. Typical add-backs: stock-based comp, store-closure charges, restructuring, impairments. TLYS has NEGATIVE Adj EBITDA in FY24 (operating loss) — EV/EBITDA shown as blank for TLYS and excluded from peer-set stats."),
        ("Loss-makers handling", "P/E is shown as blank for any peer with negative LTM EPS (CATO, CTRN, TLYS, GCO). EV/EBITDA blank for any peer with negative EBITDA (TLYS only). Excluding these cleanly from stats prevents distortion — MEDIAN and QUARTILE formulas skip empty cells."),
        ("Enterprise Value", "EV = Market Cap + Total Debt - Cash & ST Investments. Most peers are NET CASH (DXLG, BKE, CATO, CTRN, TLYS, ZUMZ) — a defining feature of conservative specialty-retail balance sheets. GCO, DBI, CURV carry meaningful net debt (CURV's reflects 2021 LBO term loan)."),
        ("Market data caveat", "Share prices in the Valuation block are illustrative placeholders (~May 22, 2026 estimates from training recall) and were NOT refreshed from a live source. REPLACE with verified pricing before relying on the multiples. Diluted share counts per each company's most recent 10-Q cover page."),
        ("Comparability caveats", "Scale difference is large: DXLG revenue $467M vs peer-set median ~$900M and DBI/GCO >$2B. Business-model heterogeneity: DXLG is a category-specialist (men's big & tall); GCO/DBI/BKE/ZUMZ are footwear or general specialty; CATO/CTRN serve different demographics. The peer set is the best LIQUID PUBLIC comp set available — not a tight pure-play group. Most direct historical comp (Men's Wearhouse / Tailored Brands) went private after 2020 bankruptcy."),
        ("Color convention", "Blue text = hardcoded input from a filing. Black text = formula. Light-grey rows = statistics (Max / 75th / Median / 25th / Min). Bolded row = DXLG (subject company). Hover over any blue cell to see its source citation."),
        ("Central question", "Is DXLG fairly valued vs specialty apparel peers, given its category-leader position, net-cash balance sheet, and structurally declining revenue base? Key tensions: (a) DXLG screens cheap on EV/EBITDA and EV/Revenue but so does the whole peer set (consumer-discretionary derate); (b) DXLG's margins are TOP-QUARTILE on EBITDA-margin despite micro-cap scale, supporting a quality premium narrative; (c) revenue declines and traffic risk could compress multiples further."),
    ]

    for i, (label, text) in enumerate(notes):
        r = NOTES_ROW + 1 + i * 2
        c = ws.cell(row=r, column=1, value=label)
        style_cell(c, bold=True, align="left", size=11)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=len(VAL_COLS))
        c2 = ws.cell(row=r, column=2, value=text)
        style_cell(c2, align="left", size=10)
        c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 45

    ws.freeze_panes = "A4"

    import os
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
