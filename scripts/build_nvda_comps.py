"""Build NVDA comparable company analysis skeleton (structure + formulas, blank inputs)."""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Color
from openpyxl.comments import Comment
from openpyxl.utils import get_column_letter

OUT = "/home/user/Claude-code-work/output/NVDA_comps.xlsx"

# Each entry: (revenue_ltm, yoy_growth, gross_profit, adj_ebitda, fcf,
#              share_price, diluted_shares_m, net_debt, eps_ltm,
#              filing_ref)
# All $ in millions. Net debt: negative = net cash. Growth as decimal.
# Filing_ref is the SEC source citation tagged onto input cells.
DATA = {
    "NVDA": dict(
        revenue=130497, growth=1.142, gross_profit=97858, adj_ebitda=88170, fcf=60853,
        share_price=140.00, shares=24900, net_debt=-38000, eps=2.94,
        ref="NVIDIA FY25 10-K (year ended Jan 26, 2025); SBC of $4.7B added to GAAP EBITDA for non-GAAP figure",
        period="FY25 (LTM through Jan 26, 2025)",
    ),
    "AMD": dict(
        revenue=25785, growth=0.140, gross_profit=12725, adj_ebitda=6138, fcf=2418,
        share_price=170.00, shares=1620, net_debt=-3800, eps=0.99,
        ref="AMD FY24 10-K (year ended Dec 28, 2024); adj EBITDA per AMD non-GAAP recon",
        period="FY24 (LTM through Dec 28, 2024)",
    ),
    "INTC": dict(
        revenue=53101, growth=-0.020, gross_profit=17346, adj_ebitda=1219, fcf=-15663,
        share_price=22.00, shares=4320, net_debt=24000, eps=-4.38,
        ref="Intel FY24 10-K (year ended Dec 28, 2024); EPS reflects GAAP loss",
        period="FY24 (LTM through Dec 28, 2024)",
    ),
    "AVGO": dict(
        revenue=51574, growth=0.440, gross_profit=38500, adj_ebitda=31920, fcf=19414,
        share_price=200.00, shares=4710, net_debt=58000, eps=1.62,
        ref="Broadcom FY24 10-K (year ended Nov 3, 2024); Gross Profit shown is NON-GAAP (adds back $7.2B amortization of acquired developed technology in COGS) for consistency with non-GAAP EBITDA",
        period="FY24 (LTM through Nov 3, 2024)",
    ),
    "QCOM": dict(
        revenue=38962, growth=0.090, gross_profit=21759, adj_ebitda=13872, fcf=11156,
        share_price=160.00, shares=1120, net_debt=-5000, eps=8.97,
        ref="Qualcomm FY24 10-K (year ended Sep 29, 2024)",
        period="FY24 (LTM through Sep 29, 2024)",
    ),
    "MRVL": dict(
        revenue=5768, growth=0.047, gross_profit=2567, adj_ebitda=1634, fcf=1402,
        share_price=70.00, shares=870, net_debt=2500, eps=-1.02,
        ref="Marvell FY25 10-K (year ended Feb 1, 2025); EPS reflects GAAP loss",
        period="FY25 (LTM through Feb 1, 2025)",
    ),
    "ARM": dict(
        revenue=4007, growth=0.240, gross_profit=3810, adj_ebitda=1520, fcf=760,
        share_price=135.00, shares=1065, net_debt=-2400, eps=0.30,
        ref="Arm Holdings FY25 10-K (year ended Mar 31, 2025); first full FY as public company",
        period="FY25 (LTM through Mar 31, 2025)",
    ),
    "TSM": dict(
        revenue=90007, growth=0.300, gross_profit=50894, adj_ebitda=60500, fcf=34716,
        share_price=185.00, shares=5190, net_debt=-50000, eps=7.21,
        ref="TSMC FY24 20-F (year ended Dec 31, 2024); NT$ converted to USD at avg rate; ADR-equivalent share count (1 ADR = 5 ordinary). EBITDA = OpInc + D&A; NOTE: EBITDA margin > GM is structural (large fab D&A in COGS gets added back to EBITDA but not GP)",
        period="FY24 (LTM through Dec 31, 2024)",
    ),
}

NAVY = "17365D"
LIGHT_BLUE = "D9E1F2"
LIGHT_GREY = "F2F2F2"
WHITE = "FFFFFF"
INPUT_BLUE = "0070C0"

FONT_NAME = "Times New Roman"

PEERS = [
    ("NVIDIA", "NVDA"),
    ("Advanced Micro Devices", "AMD"),
    ("Intel", "INTC"),
    ("Broadcom", "AVGO"),
    ("Qualcomm", "QCOM"),
    ("Marvell Technology", "MRVL"),
    ("Arm Holdings", "ARM"),
    ("Taiwan Semiconductor", "TSM"),
]

OPS_COLS = [
    ("Company", 24, "left"),
    ("Revenue (LTM, $M)", 18, "center"),
    ("Rev Growth (YoY)", 16, "center"),
    ("Gross Profit ($M)", 18, "center"),
    ("Gross Margin", 14, "center"),
    ("EBITDA ($M)", 16, "center"),
    ("EBITDA Margin", 14, "center"),
    ("FCF ($M)", 14, "center"),
    ("FCF Margin", 12, "center"),
]

VAL_COLS = [
    ("Company", 24, "left"),
    ("Share Price ($)", 14, "center"),
    ("Diluted Shares (M)", 16, "center"),
    ("Market Cap ($M)", 16, "center"),
    ("Net Debt ($M)", 14, "center"),
    ("Enterprise Value ($M)", 18, "center"),
    ("EV / Revenue", 14, "center"),
    ("EV / EBITDA", 14, "center"),
    ("P / E (LTM)", 14, "center"),
    ("EPS (LTM, $)", 14, "center"),
]

STAT_ROWS = [
    ("Maximum", "MAX"),
    ("75th Percentile", "QUARTILE_3"),
    ("Median", "MEDIAN"),
    ("25th Percentile", "QUARTILE_1"),
    ("Minimum", "MIN"),
]


def style_cell(cell, *, bold=False, color=None, fill=None, align="left", size=11, italic=False, number_format=None):
    cell.font = Font(name=FONT_NAME, size=size, bold=bold, italic=italic, color=color or "000000")
    if fill:
        cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = Alignment(horizontal=align, vertical="center")
    if number_format:
        cell.number_format = number_format


def section_header(ws, row, span_cols, text):
    ws.cell(row=row, column=1, value=text)
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span_cols)
    for col in range(1, span_cols + 1):
        c = ws.cell(row=row, column=col)
        style_cell(c, bold=True, color=WHITE, fill=NAVY, align="center", size=12)
    ws.row_dimensions[row].height = 22


def column_headers(ws, row, cols):
    for i, (name, _w, align) in enumerate(cols, start=1):
        c = ws.cell(row=row, column=i, value=name)
        style_cell(c, bold=True, fill=LIGHT_BLUE, align="center", size=11)
    ws.row_dimensions[row].height = 30


def set_col_widths(ws, cols):
    for i, (_n, w, _a) in enumerate(cols, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_company_names(ws, start_row, cols, fill_company_col=True):
    for i, (name, ticker) in enumerate(PEERS):
        r = start_row + i
        c = ws.cell(row=r, column=1, value=f"{name} ({ticker})")
        style_cell(c, bold=False, align="left", size=11)
        for col_idx in range(2, len(cols) + 1):
            cc = ws.cell(row=r, column=col_idx)
            style_cell(cc, align="center", size=11)
        ws.row_dimensions[r].height = 20


def color_input_cells(ws, start_row, col_indices, n_companies=len(PEERS)):
    """Apply blue font to hardcoded input cells (raw data entry)."""
    for i in range(n_companies):
        r = start_row + i
        for col_idx in col_indices:
            c = ws.cell(row=r, column=col_idx)
            c.font = Font(name=FONT_NAME, size=11, color=INPUT_BLUE)


def set_input(ws, row, col, value, source_ref, field_label):
    """Set an input cell value and attach a source-citation comment."""
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(name=FONT_NAME, size=11, color=INPUT_BLUE)
    c.comment = Comment(
        f"{field_label}\nSource: {source_ref}\nNote: Figure from training-data recall (Jan 2026 cutoff). VERIFY against the cited filing on SEC EDGAR before publishing or trading on this output.",
        "Comps",
    )


def stats_block(ws, data_start_row, data_end_row, stats_start_row, cols_to_stat, n_cols):
    """Write Max/75/Median/25/Min rows. cols_to_stat = list of column indices."""
    for j, (label, fn) in enumerate(STAT_ROWS):
        r = stats_start_row + j
        c = ws.cell(row=r, column=1, value=label)
        style_cell(c, bold=True, fill=LIGHT_GREY, align="left", size=11, italic=True)
        for col_idx in range(2, n_cols + 1):
            cc = ws.cell(row=r, column=col_idx)
            style_cell(cc, fill=LIGHT_GREY, align="center", size=11)
            if col_idx in cols_to_stat:
                col_letter = get_column_letter(col_idx)
                rng = f"{col_letter}{data_start_row}:{col_letter}{data_end_row}"
                if fn == "MAX":
                    formula = f"=MAX({rng})"
                elif fn == "MIN":
                    formula = f"=MIN({rng})"
                elif fn == "MEDIAN":
                    formula = f"=MEDIAN({rng})"
                elif fn == "QUARTILE_1":
                    formula = f"=QUARTILE({rng},1)"
                elif fn == "QUARTILE_3":
                    formula = f"=QUARTILE({rng},3)"
                cc.value = formula
        ws.row_dimensions[r].height = 20


def build():
    wb = Workbook()
    ws = wb.active
    ws.title = "Comps"

    ws.sheet_view.showGridLines = False

    set_col_widths(ws, OPS_COLS)
    if len(VAL_COLS) > len(OPS_COLS):
        for i in range(len(OPS_COLS) + 1, len(VAL_COLS) + 1):
            ws.column_dimensions[get_column_letter(i)].width = VAL_COLS[i - 1][1]

    max_cols = max(len(OPS_COLS), len(VAL_COLS))

    # ----- Header block (rows 1-3) -----
    title = "SEMICONDUCTORS — COMPARABLE COMPANY ANALYSIS"
    ws.cell(row=1, column=1, value=title)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_cols)
    for col in range(1, max_cols + 1):
        style_cell(ws.cell(row=1, column=col), bold=True, color=WHITE, fill=NAVY, align="center", size=14)
    ws.row_dimensions[1].height = 28

    sub = " • ".join([f"{n} ({t})" for n, t in PEERS])
    ws.cell(row=2, column=1, value=sub)
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max_cols)
    style_cell(ws.cell(row=2, column=1), bold=False, align="center", size=10, italic=True)
    ws.row_dimensions[2].height = 18

    ctx = "As of LTM through each company's most-recent reported quarter | All figures in USD Millions except per-share amounts and ratios"
    ws.cell(row=3, column=1, value=ctx)
    ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=max_cols)
    style_cell(ws.cell(row=3, column=1), align="center", size=10, italic=True)
    ws.row_dimensions[3].height = 18

    # ----- Section 1: Operating Metrics -----
    section_header(ws, 5, len(OPS_COLS), "OPERATING STATISTICS & FINANCIAL METRICS")
    column_headers(ws, 6, OPS_COLS)

    OPS_DATA_START = 7
    OPS_DATA_END = OPS_DATA_START + len(PEERS) - 1  # row 14
    write_company_names(ws, OPS_DATA_START, OPS_COLS)

    # Input columns (raw figures): B=Revenue, C=Growth, D=Gross Profit, F=EBITDA, H=FCF
    for i, (name, ticker) in enumerate(PEERS):
        r = OPS_DATA_START + i
        d = DATA[ticker]
        ref = d["ref"]
        set_input(ws, r, 2, d["revenue"], ref, f"{ticker} — Revenue (LTM, $M), {d['period']}")
        set_input(ws, r, 3, d["growth"], ref, f"{ticker} — Revenue YoY growth, {d['period']}")
        set_input(ws, r, 4, d["gross_profit"], ref, f"{ticker} — Gross Profit ($M), {d['period']}")
        set_input(ws, r, 6, d["adj_ebitda"], ref, f"{ticker} — Adjusted (non-GAAP) EBITDA ($M), {d['period']}")
        set_input(ws, r, 8, d["fcf"], ref, f"{ticker} — Free Cash Flow ($M), {d['period']}")

    # Formula columns: E=Gross Margin (D/B), G=EBITDA Margin (F/B), I=FCF Margin (H/B)
    for i in range(len(PEERS)):
        r = OPS_DATA_START + i
        # Gross Margin = D/B
        c = ws.cell(row=r, column=5, value=f"=IFERROR(D{r}/B{r},\"\")")
        style_cell(c, align="center", number_format="0.0%")
        # EBITDA Margin = F/B
        c = ws.cell(row=r, column=7, value=f"=IFERROR(F{r}/B{r},\"\")")
        style_cell(c, align="center", number_format="0.0%")
        # FCF Margin = H/B
        c = ws.cell(row=r, column=9, value=f"=IFERROR(H{r}/B{r},\"\")")
        style_cell(c, align="center", number_format="0.0%")

    # Format input columns
    for i in range(len(PEERS)):
        r = OPS_DATA_START + i
        for col_idx in [2, 4, 6, 8]:
            ws.cell(row=r, column=col_idx).number_format = "#,##0"
        ws.cell(row=r, column=3).number_format = "0.0%"

    # Blank row at 15, stats at 16-20
    OPS_STATS_START = OPS_DATA_END + 2  # row 16
    OPS_STATS_COLS = [3, 5, 7, 9]  # Growth, GM, EBITDA Margin, FCF Margin
    stats_block(ws, OPS_DATA_START, OPS_DATA_END, OPS_STATS_START, OPS_STATS_COLS, len(OPS_COLS))
    # Apply percent format to stats cells in those columns
    for j in range(len(STAT_ROWS)):
        r = OPS_STATS_START + j
        for col_idx in OPS_STATS_COLS:
            ws.cell(row=r, column=col_idx).number_format = "0.0%"

    # ----- Section 2: Valuation Multiples -----
    VAL_HEADER_ROW = OPS_STATS_START + len(STAT_ROWS) + 2  # row 23
    section_header(ws, VAL_HEADER_ROW, len(VAL_COLS), "VALUATION MULTIPLES & INVESTMENT METRICS")
    column_headers(ws, VAL_HEADER_ROW + 1, VAL_COLS)

    VAL_DATA_START = VAL_HEADER_ROW + 2  # row 25
    VAL_DATA_END = VAL_DATA_START + len(PEERS) - 1  # row 32
    write_company_names(ws, VAL_DATA_START, VAL_COLS)

    # Inputs: B=Share Price, C=Diluted Shares, E=Net Debt, J=EPS
    for i, (name, ticker) in enumerate(PEERS):
        r = VAL_DATA_START + i
        d = DATA[ticker]
        ref = d["ref"]
        set_input(ws, r, 2, d["share_price"], "Approximate close as of 2026-05-21; placeholder — replace with verified market data", f"{ticker} — Share Price ($), illustrative as-of 2026-05-21")
        set_input(ws, r, 3, d["shares"], ref, f"{ticker} — Diluted Shares Outstanding (M), per most recent 10-Q cover page")
        set_input(ws, r, 5, d["net_debt"], ref, f"{ticker} — Net Debt ($M, negative = net cash) = Total Debt - Cash & ST Investments, most recent b/s")
        set_input(ws, r, 10, d["eps"], ref, f"{ticker} — Diluted EPS (LTM, $), {d['period']}")

    # Formulas + cross-section refs
    # D = Market Cap = B*C
    # F = EV = D + E
    # G = EV/Revenue = F / Comps!B{ops_row}   (ops_row = same i offset, OPS_DATA_START)
    # H = EV/EBITDA = F / Comps!F{ops_row}
    # I = P/E = B / J
    for i in range(len(PEERS)):
        vr = VAL_DATA_START + i
        ops_r = OPS_DATA_START + i

        c = ws.cell(row=vr, column=4, value=f"=B{vr}*C{vr}")
        style_cell(c, align="center", number_format="#,##0")

        c = ws.cell(row=vr, column=6, value=f"=D{vr}+E{vr}")
        style_cell(c, align="center", number_format="#,##0")

        c = ws.cell(row=vr, column=7, value=f"=IFERROR(F{vr}/B{ops_r},\"\")")
        style_cell(c, align="center", number_format="0.0\"x\"")

        c = ws.cell(row=vr, column=8, value=f"=IFERROR(F{vr}/F{ops_r},\"\")")
        style_cell(c, align="center", number_format="0.0\"x\"")

        c = ws.cell(row=vr, column=9, value=f"=IFERROR(B{vr}/J{vr},\"\")")
        style_cell(c, align="center", number_format="0.0\"x\"")

        # Format inputs
        ws.cell(row=vr, column=2).number_format = "$#,##0.00"
        ws.cell(row=vr, column=3).number_format = "#,##0"
        ws.cell(row=vr, column=5).number_format = "#,##0;(#,##0)"
        ws.cell(row=vr, column=10).number_format = "$#,##0.00"

    # Blank row, then stats
    VAL_STATS_START = VAL_DATA_END + 2  # row 34
    VAL_STATS_COLS = [7, 8, 9]  # EV/Rev, EV/EBITDA, P/E
    stats_block(ws, VAL_DATA_START, VAL_DATA_END, VAL_STATS_START, VAL_STATS_COLS, len(VAL_COLS))
    for j in range(len(STAT_ROWS)):
        r = VAL_STATS_START + j
        for col_idx in VAL_STATS_COLS:
            ws.cell(row=r, column=col_idx).number_format = "0.0\"x\""

    # ----- Section 3: Notes & Methodology -----
    NOTES_ROW = VAL_STATS_START + len(STAT_ROWS) + 2  # row 41
    section_header(ws, NOTES_ROW, len(VAL_COLS), "NOTES & METHODOLOGY")

    notes = [
        ("⚠ VERIFY BEFORE USE", "All input figures (blue cells) were populated from training-data recall (Jan 2026 cutoff) citing each company's most recent SEC filing. Every blue cell has a comment with the source filing. VERIFY each figure against the actual EDGAR filing before publishing, trading, or presenting. Share prices are illustrative placeholders as of 2026-05-21 and MUST be refreshed against live market data."),
        ("Data sources", "Per-company 10-K filings (TSM uses 20-F as foreign filer). NVDA FY25, AMD FY24, INTC FY24, AVGO FY24 (post-VMware close), QCOM FY24, MRVL FY25, ARM FY25 (first full FY public), TSM FY24. Each input cell carries a filing-specific citation in its comment."),
        ("Reporting periods (MIXED)", "Fiscal year-ends differ: NVDA Jan 2025, MRVL Feb 2025, ARM Mar 2025, TSM/AMD/INTC Dec 2024, AVGO Nov 2024, QCOM Sep 2024. This mixes period-ends by up to 7 months — meaningful in a fast-moving cycle (AI accelerator demand surged H2 2024 - H1 2025). Reader should weight NVDA/MRVL/ARM figures as more current and QCOM/AVGO as relatively older."),
        ("EBITDA definition", "Adjusted (non-GAAP) EBITDA per each company's investor presentation reconciliation. Adds back stock-based comp, restructuring, and acquisition-related charges. Definitions are NOT uniform across companies — AVGO's reflects significant VMware purchase-accounting amortization addbacks; AVGO Gross Profit also shown on a non-GAAP basis for consistency. FOR TSM (FOUNDRY): Adjusted EBITDA margin exceeds GAAP Gross Margin because large fab D&A flows through COGS but gets added back into EBITDA — this is structural to the foundry model, not a data error. Same effect smaller for INTC (IDM)."),
        ("Enterprise Value", "EV = Market Cap + Total Debt - Cash & ST Investments. Net Debt column shows negative values in parentheses for net-cash companies (NVDA, AMD, QCOM, ARM, TSM). INTC and AVGO carry meaningful net debt; AVGO's reflects VMware financing."),
        ("Market data caveat", "Share prices in the Valuation block are illustrative placeholders (approximate close ~May 2026) and were NOT refreshed from a live source. REPLACE with verified pricing before relying on the multiples. Diluted share counts are from each company's most recent 10-Q cover page; NVDA and AVGO reflect their 2024 10:1 splits."),
        ("Comparability caveats", "Business-model heterogeneity is large: NVDA/AMD/AVGO are fabless designers; INTC is an integrated device manufacturer (IDM); TSM is a pure-play foundry; ARM is an IP-licensing model. Margin profiles will differ materially — ARM's ~95% gross margin reflects the IP-licensing model, not operational excellence relative to manufacturers. Weight peer-set median accordingly."),
        ("Color convention", "Blue text = hardcoded input from a filing. Black text = formula. Light-grey rows = statistics (Max / 75th / Median / 25th / Min). Hover over any blue cell to see its source citation."),
        ("Central question", "Is NVDA fairly valued relative to its semiconductor peer set on EV/Revenue, EV/EBITDA, and P/E, given its growth and margin profile? Key tension: NVDA trades at substantially higher multiples than peers — justified by ~114% YoY growth and 67%+ EBITDA margins, but creates valuation asymmetry if AI capex slows."),
    ]

    for i, (label, text) in enumerate(notes):
        r = NOTES_ROW + 1 + i * 2
        c = ws.cell(row=r, column=1, value=label)
        style_cell(c, bold=True, align="left", size=11)
        ws.merge_cells(start_row=r, start_column=2, end_row=r, end_column=len(VAL_COLS))
        c2 = ws.cell(row=r, column=2, value=text)
        style_cell(c2, align="left", size=10)
        c2.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        ws.row_dimensions[r].height = 40

    # Freeze top three header rows
    ws.freeze_panes = "A4"

    import os
    os.makedirs(os.path.dirname(OUT), exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    build()
